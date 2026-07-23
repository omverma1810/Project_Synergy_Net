import io
from pathlib import Path
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from analysis.insights import derive_risk_flags, risk_summary, build_finance_snapshot


class PDFReportGenerator:
    def __init__(self, analysis):
        self.analysis = analysis
        self.results = list(analysis.results.all().select_related('territory').order_by('rank'))
        self.project = analysis.project

    def _build_context(self):
        # Annotate each result with its derived risk profile for the template.
        enriched = []
        for r in self.results:
            r.risk_flags = derive_risk_flags(r)
            r.risk_level = risk_summary(r)
            enriched.append(r)

        snapshot = build_finance_snapshot(self.analysis)
        producer = getattr(self.project, 'producer', None)
        company = getattr(producer, 'company_name', '') if producer else ''

        # Build a territory-id → territory detail map for the top results, so the
        # template can render film commission links and government support without
        # extra DB queries per row.
        top_territory_ids = [r.territory_id for r in enriched[:5]]
        from territories.models import Territory
        territory_detail_map = {
            t.id: t
            for t in Territory.objects.filter(id__in=top_territory_ids)
        }
        for r in enriched:
            r.territory_detail = territory_detail_map.get(r.territory_id)

        return {
            'project': self.project,
            'analysis': self.analysis,
            'results': enriched,
            'top_results': enriched[:5],
            'winner': enriched[0] if enriched else None,
            'snapshot': snapshot,
            'company': company,
            'producer': producer,
            'generated_at': timezone.now(),
            'year': timezone.now().year,
            'logo_data': self._load_logo(),
        }

    @staticmethod
    def _load_logo():
        """Base64-encode the Synergy Media Labs logo for inline embedding."""
        import base64
        logo_path = Path(__file__).resolve().parent / 'static' / 'logo.jpg'
        try:
            return base64.b64encode(logo_path.read_bytes()).decode()
        except OSError:
            return ''

    def generate(self):
        html_string = render_to_string('reports/analysis_report.html', self._build_context())
        try:
            from weasyprint import HTML
        except ImportError as exc:
            raise RuntimeError(
                'PDF report generation requires WeasyPrint and its native system dependencies.'
            ) from exc
        pdf = HTML(string=html_string, base_url=str(settings.BASE_DIR)).write_pdf()
        return pdf


class ExcelReportGenerator:
    def __init__(self, analysis):
        self.analysis = analysis
        self.results = analysis.results.all().select_related('territory')
        self.project = analysis.project

    def generate(self):
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Summary"
        ws.append(["Project", self.project.title])
        ws.append(["Analysis Date", self.analysis.created_at.strftime("%Y-%m-%d %H:%M")])
        ws.append(["Total Budget", f"{self.project.total_budget} {self.project.currency}"])
        ws.append([])
        ws.append([
            "Rank", "Territory", "Rebate %", "Qualified Spend",
            "Estimated Rebate", "Logistics Premium", "Net Benefit", "Confidence",
        ])

        for result in self.results:
            ws.append([
                result.rank,
                result.territory.name,
                f"{result.territory.base_percentage}%",
                float(result.qualified_spend_total),
                float(result.estimated_rebate),
                float(result.logistics_premium),
                float(result.net_benefit),
                f"{result.confidence_score * 100}%",
            ])

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        ws2 = wb.create_sheet("Details")
        ws2.append(["Territory", "Category", "Total Amount", "Qualified Amount"])
        for result in self.results:
            for cat, vals in result.details.get('category_breakdown', {}).items():
                ws2.append([
                    result.territory.name,
                    cat,
                    vals['total'],
                    vals['qualified'],
                ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
